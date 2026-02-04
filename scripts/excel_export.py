"""
Excel file generation with formatting.
Exports Twitter data to styled Excel spreadsheet.
"""

import argparse
import json
from datetime import datetime
from pathlib import Path
from typing import Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config import Config
from utils import setup_logger


logger = setup_logger(__name__, Config.LOG_LEVEL)


class ExcelExporter:
    """Excel file generator with formatting"""

    # Color scheme
    HEADER_FILL = PatternFill(start_color="C2185B", end_color="C2185B", fill_type="solid")
    SUMMARY_FILL = PatternFill(start_color="F8BBD0", end_color="F8BBD0", fill_type="solid")

    # Fonts
    HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    SUMMARY_FONT = Font(name="Arial", size=10, bold=True, color="1D1D1F")
    REGULAR_FONT = Font(name="Arial", size=10, color="1D1D1F")
    LINK_FONT = Font(name="Arial", size=10, color="0563C1", underline="single")

    # Borders
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Column widths
    COLUMN_WIDTHS = {
        'A': 12,  # Date
        'B': 15,  # Username
        'C': 18,  # Display Name
        'D': 50,  # AI Summary
        'E': 60,  # Tweet Content
        'F': 40,  # URL
        'G': 10,  # Likes
        'H': 10,  # Retweets
        'I': 18,  # Timestamp
    }

    def __init__(self):
        """Initialize Excel exporter"""
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Twitter Summary"
        self.logger = logger
        self.current_row = 1

    def _write_headers(self):
        """Write column headers with styling"""
        headers = [
            "日期",
            "用户名",
            "显示名称",
            "AI摘要",
            "推文内容",
            "原始链接",
            "点赞数",
            "转发数",
            "发布时间"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.THIN_BORDER

        self.current_row = 2
        self.logger.debug("Headers written")

    def _write_data(self, data: Dict):
        """
        Write account data grouped by user.

        Args:
            data: Data dictionary from JSON file
        """
        scrape_date = data.get('date', datetime.now().strftime('%Y-%m-%d'))
        accounts = data.get('accounts', [])

        for account in accounts:
            username = account['username']
            display_name = account['display_name']
            ai_summary = account.get('ai_summary', 'AI摘要暂时不可用')
            tweets = account['tweets']

            if not tweets:
                # If no tweets, still write a row for the account
                self._write_account_row(
                    scrape_date, username, display_name, ai_summary,
                    "暂无推文", "", 0, 0, ""
                )
                continue

            # Write first tweet row with AI summary
            first_tweet = tweets[0]
            self._write_account_row(
                scrape_date, username, display_name, ai_summary,
                first_tweet['text'],
                first_tweet['url'],
                first_tweet['likes'],
                first_tweet['retweets'],
                first_tweet['created_at'],
                is_summary_row=True
            )

            # Write remaining tweets without AI summary
            for tweet in tweets[1:]:
                self._write_account_row(
                    scrape_date, username, display_name, "",
                    tweet['text'],
                    tweet['url'],
                    tweet['likes'],
                    tweet['retweets'],
                    tweet['created_at'],
                    is_summary_row=False
                )

        self.logger.info(f"Written {len(accounts)} accounts to Excel")

    def _write_account_row(
        self,
        date: str,
        username: str,
        display_name: str,
        ai_summary: str,
        tweet_text: str,
        tweet_url: str,
        likes: int,
        retweets: int,
        created_at: str,
        is_summary_row: bool = False
    ):
        """Write a single row of data"""
        row = self.current_row

        # Date
        cell = self.ws.cell(row=row, column=1, value=date)
        cell.font = self.REGULAR_FONT
        cell.alignment = Alignment(horizontal='center', vertical='top')
        cell.border = self.THIN_BORDER

        # Username
        cell = self.ws.cell(row=row, column=2, value=f"@{username}")
        cell.font = self.REGULAR_FONT
        cell.alignment = Alignment(horizontal='left', vertical='top')
        cell.border = self.THIN_BORDER

        # Display Name
        cell = self.ws.cell(row=row, column=3, value=display_name)
        cell.font = self.REGULAR_FONT
        cell.alignment = Alignment(horizontal='left', vertical='top')
        cell.border = self.THIN_BORDER

        # AI Summary (only in first row per account)
        cell = self.ws.cell(row=row, column=4, value=ai_summary)
        if is_summary_row and ai_summary:
            cell.font = self.SUMMARY_FONT
            cell.fill = self.SUMMARY_FILL
        else:
            cell.font = self.REGULAR_FONT
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.border = self.THIN_BORDER

        # Tweet Content
        cell = self.ws.cell(row=row, column=5, value=tweet_text)
        cell.font = self.REGULAR_FONT
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.border = self.THIN_BORDER

        # Tweet URL (hyperlink)
        cell = self.ws.cell(row=row, column=6, value=tweet_url)
        if tweet_url:
            cell.hyperlink = tweet_url
            cell.font = self.LINK_FONT
        else:
            cell.font = self.REGULAR_FONT
        cell.alignment = Alignment(horizontal='left', vertical='top')
        cell.border = self.THIN_BORDER

        # Likes
        cell = self.ws.cell(row=row, column=7, value=likes)
        cell.font = self.REGULAR_FONT
        cell.alignment = Alignment(horizontal='right', vertical='top')
        cell.number_format = '#,##0'
        cell.border = self.THIN_BORDER

        # Retweets
        cell = self.ws.cell(row=row, column=8, value=retweets)
        cell.font = self.REGULAR_FONT
        cell.alignment = Alignment(horizontal='right', vertical='top')
        cell.number_format = '#,##0'
        cell.border = self.THIN_BORDER

        # Created At
        cell = self.ws.cell(row=row, column=9, value=created_at)
        cell.font = self.REGULAR_FONT
        cell.alignment = Alignment(horizontal='center', vertical='top')
        cell.border = self.THIN_BORDER

        # Set row height for better readability
        self.ws.row_dimensions[row].height = 30

        self.current_row += 1

    def _apply_formatting(self):
        """Apply cell formatting and column widths"""
        # Set column widths
        for col_letter, width in self.COLUMN_WIDTHS.items():
            self.ws.column_dimensions[col_letter].width = width

        # Freeze header row
        self.ws.freeze_panes = 'A2'

        self.logger.debug("Formatting applied")

    def export(self, data: Dict, output_path: Path):
        """
        Main export function.

        Args:
            data: Data dictionary from JSON file
            output_path: Path to save Excel file
        """
        self._write_headers()
        self._write_data(data)
        self._apply_formatting()
        self._save(output_path)

    def _save(self, output_path: Path):
        """Save workbook to file"""
        try:
            self.wb.save(output_path)
            self.logger.info(f"Excel file saved to {output_path}")
        except Exception as e:
            self.logger.error(f"Failed to save Excel file: {e}")
            raise


def main():
    """Main entry point for standalone execution"""
    parser = argparse.ArgumentParser(description="Export Twitter data to Excel")
    parser.add_argument(
        "--input",
        required=True,
        help="Path to input JSON file (e.g., data/2026-02-04.json)"
    )
    args = parser.parse_args()

    input_path = Path(args.input)

    # Validate input file
    if not input_path.exists():
        logger.error(f"Input file not found: {input_path}")
        return 1

    # Ensure output directory exists
    Config.ensure_directories()

    # Determine output path
    date_str = input_path.stem  # e.g., "2026-02-04"
    output_path = Config.OUTPUT_DIR / f"{date_str}.xlsx"

    # Load data
    try:
        with open(input_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        logger.info(f"Loaded data from {input_path}")
    except Exception as e:
        logger.error(f"Failed to load input file: {e}")
        return 1

    # Export to Excel
    try:
        exporter = ExcelExporter()
        exporter.export(data, output_path)
        logger.info("Excel export completed successfully")
        return 0
    except Exception as e:
        logger.error(f"Excel export failed: {e}")
        return 1


if __name__ == "__main__":
    exit(main())
